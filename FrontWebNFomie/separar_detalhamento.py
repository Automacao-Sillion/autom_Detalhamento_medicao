# -*- coding: utf-8 -*-
"""
separar_detalhamento.py  (versao FRONT)
=======================================
Adaptacao do separar_detalhamento.py original para rodar no front Streamlit e
enviar o resultado via webhook.

Mantem a logica do script original:
  - le a aba 'Faturamento' (12 colunas A..L), periodo na linha 2, cabecalho na 5,
    dados a partir da 6;
  - agrupa por CONT ID (mantendo a ordem de aparicao);
  - gera 1 .xlsx por CONT ID com SUBTOTAL e linha "<CONT ID> Total";
  - nome: "<CONT ID> - <CLIENTE NOME> - <VALOR TOTAL>.xlsx".

Diferencas em relacao ao script de linha de comando:
  - entrada/saida em MEMORIA (bytes), sem caminhos de disco;
  - o "template" de estilo vem da PROPRIA aba 'Faturamento' do arquivo enviado
    (rotulos do periodo nas linhas 1-2, cabecalho na 5, dados na 6 e a linha
    "TOTAL (R$)" na 4 como modelo de subtotal/total), pois o arquivo padrao tem
    uma unica aba.
"""

import io
import re
from copy import copy
from collections import OrderedDict

import openpyxl

# --------------------------------------------------------------------------- #
# CONFIG
# --------------------------------------------------------------------------- #
ABA_FONTE = "Faturamento"

# Layout da ORIGEM (aba unica enviada)
LINHA_PERIODO_FONTE   = 2
LINHA_CABECALHO_FONTE = 5
LINHA_DADOS_FONTE     = 6
N_COLUNAS             = 12

COL_PROPORCIONAL = 9    # J = PROPORCIONAL MENSAL
COL_OPERACAO     = 10   # K = VALOR OPERACAO

# Layout da SAIDA (padrao "Detalhamentos PDF" gerado pelo script original)
OUT_PERIODO_LBL = 1
OUT_PERIODO_VAL = 2
OUT_CABECALHO   = 4
OUT_DADOS       = 5

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --------------------------------------------------------------------------- #
# UTILITARIOS
# --------------------------------------------------------------------------- #
def copiar_estilo(origem, destino):
    """Copia fonte, preenchimento, borda, alinhamento, formato e protecao."""
    if origem.has_style:
        destino.font          = copy(origem.font)
        destino.fill          = copy(origem.fill)
        destino.border        = copy(origem.border)
        destino.alignment     = copy(origem.alignment)
        destino.protection    = copy(origem.protection)
        destino.number_format = origem.number_format


def sanitizar_nome(texto, limite=110):
    """Remove caracteres invalidos para nome de arquivo no Windows."""
    texto = str(texto)
    texto = re.sub(r'[\\/:*?"<>|\r\n\t]+', " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip(" .")
    return texto[:limite].strip() or "SEM_NOME"


def total_medicao(linhas):
    """Total do bloco = soma PROPORCIONAL MENSAL (J) + soma VALOR OPERACAO (K)."""
    prop = sum(v[COL_PROPORCIONAL] for v in linhas
               if isinstance(v[COL_PROPORCIONAL], (int, float)))
    oper = sum(v[COL_OPERACAO] for v in linhas
               if isinstance(v[COL_OPERACAO], (int, float)))
    return prop + oper


def capturar_modelo(ws):
    """Captura os prototipos de estilo da PROPRIA aba de origem."""
    def linha_proto(linha):
        return [ws.cell(row=linha, column=c) for c in range(1, N_COLUNAS + 1)]
    return {
        "periodo_lbl": linha_proto(1),
        "periodo_val": linha_proto(2),
        "cabecalho":   linha_proto(LINHA_CABECALHO_FONTE),  # r5: titulos das colunas
        "dados":       linha_proto(LINHA_DADOS_FONTE),      # r6: 1a linha de dados
        "subtotal":    linha_proto(4),                      # r4: "TOTAL (R$)" como modelo
        "total":       linha_proto(4),
        "larguras": {col: dim.width for col, dim in ws.column_dimensions.items()
                     if dim.width is not None},
        "freeze":        ws.freeze_panes,
        "orientation":   ws.page_setup.orientation,
        "paperSize":     ws.page_setup.paperSize,
        "scale":         ws.page_setup.scale,
        "fitToWidth":    ws.page_setup.fitToWidth,
        "fitToHeight":   ws.page_setup.fitToHeight,
        "fitToPage":     bool(ws.sheet_properties.pageSetUpPr and
                              ws.sheet_properties.pageSetUpPr.fitToPage),
        "margins":       (ws.page_margins.left, ws.page_margins.right,
                          ws.page_margins.top, ws.page_margins.bottom),
        "showGridLines": ws.sheet_view.showGridLines,
        "alturaLinha":   ws.row_dimensions[LINHA_DADOS_FONTE].height or 15.0,
    }


def ler_dados_fonte(wb):
    """Le a aba 'Faturamento' e devolve (periodo, grupos, nome_cliente)."""
    ws = wb[ABA_FONTE]
    periodo = (None, None, None)
    grupos = OrderedDict()
    nome_cliente = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == LINHA_PERIODO_FONTE:
            periodo = (row[5], row[6], row[7])   # F2 / G2 / H2
        if i < LINHA_DADOS_FONTE or row is None:
            continue
        cont_id = row[0]
        if cont_id is None or cont_id == "":
            continue
        valores = tuple(row[c] if c < len(row) else None for c in range(N_COLUNAS))
        grupos.setdefault(cont_id, []).append(valores)
        if cont_id not in nome_cliente:
            nome_cliente[cont_id] = row[2]       # C = CLIENTE NOME
    return periodo, grupos, nome_cliente


# --------------------------------------------------------------------------- #
# CONSTRUCAO DE UM BLOCO (1 .xlsx por CONT ID)
# --------------------------------------------------------------------------- #
def construir_bloco(cont_id, linhas, periodo, modelo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalhamentos PDF"

    n = len(linhas)
    linha_subtotal = OUT_DADOS + n
    linha_total = linha_subtotal + 1

    for col, largura in modelo["larguras"].items():
        ws.column_dimensions[col].width = largura

    # linha 1: rotulos do periodo
    for c in range(1, N_COLUNAS + 1):
        proto = modelo["periodo_lbl"][c - 1]
        copiar_estilo(proto, ws.cell(row=1, column=c, value=proto.value))

    # linha 2: PERIODO + datas da origem
    for c in range(1, N_COLUNAS + 1):
        proto = modelo["periodo_val"][c - 1]
        cel = ws.cell(row=2, column=c)
        copiar_estilo(proto, cel)
        if c == 1:
            cel.value = proto.value
        elif c == 6:
            cel.value = periodo[0]
        elif c == 7:
            cel.value = periodo[1]
        elif c == 8:
            cel.value = periodo[2]

    # linha 4: cabecalho das colunas
    for c in range(1, N_COLUNAS + 1):
        proto = modelo["cabecalho"][c - 1]
        copiar_estilo(proto, ws.cell(row=OUT_CABECALHO, column=c, value=proto.value))

    # linhas de dados
    for idx, valores in enumerate(linhas):
        r = OUT_DADOS + idx
        for c in range(1, N_COLUNAS + 1):
            proto = modelo["dados"][c - 1]
            copiar_estilo(proto, ws.cell(row=r, column=c, value=valores[c - 1]))

    # linha de SUBTOTAL
    for c in range(1, N_COLUNAS + 1):
        proto = modelo["subtotal"][c - 1]
        copiar_estilo(proto, ws.cell(row=linha_subtotal, column=c))
    ws.cell(row=linha_subtotal, column=10).value = \
        f"=SUBTOTAL(9,J{OUT_DADOS}:J{linha_subtotal - 1})"
    ws.cell(row=linha_subtotal, column=11).value = \
        f"=SUBTOTAL(9,K{OUT_DADOS}:K{linha_subtotal - 1})"

    # linha de TOTAL
    for c in range(1, N_COLUNAS + 1):
        proto = modelo["total"][c - 1]
        copiar_estilo(proto, ws.cell(row=linha_total, column=c))
    ws.cell(row=linha_total, column=1).value  = f"{cont_id} Total"
    ws.cell(row=linha_total, column=10).value = f"=J{linha_subtotal}+K{linha_subtotal}"

    # alturas / visao / impressao
    for r in range(1, linha_total + 1):
        ws.row_dimensions[r].height = modelo["alturaLinha"]
    ws.freeze_panes = modelo["freeze"]
    ws.sheet_view.showGridLines = modelo["showGridLines"]
    ws.page_setup.orientation = modelo["orientation"]
    ws.page_setup.paperSize   = modelo["paperSize"]
    ws.page_setup.scale       = modelo["scale"]
    ws.page_setup.fitToWidth  = modelo["fitToWidth"]
    ws.page_setup.fitToHeight = modelo["fitToHeight"]
    if modelo["fitToPage"]:
        ws.sheet_properties.pageSetUpPr = \
            openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins.left, ws.page_margins.right, ws.page_margins.top, ws.page_margins.bottom = \
        modelo["margins"]
    ws.print_area = f"A1:L{linha_total}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# --------------------------------------------------------------------------- #
# API PUBLICA
# --------------------------------------------------------------------------- #
def gerar_blocos(file_bytes):
    """
    Processa o arquivo enviado e gera 1 .xlsx por CONT ID (em memoria).
    Retorna:
    {
      "periodo": (data_inicial, data_final, qtd_dias),
      "blocos": [
        {cont_id, cliente_nome, total, qtd_veiculos, filename, xlsx_bytes}, ...
      ]
    }
    """
    wb_valores = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_estilos = openpyxl.load_workbook(io.BytesIO(file_bytes))

    periodo, grupos, nome_cliente = ler_dados_fonte(wb_valores)
    modelo = capturar_modelo(wb_estilos[ABA_FONTE])

    blocos = []
    usados = set()
    for cont_id, linhas in grupos.items():
        xls = construir_bloco(cont_id, linhas, periodo, modelo)
        total = total_medicao(linhas)
        cliente = sanitizar_nome(nome_cliente.get(cont_id) or "", limite=80)
        base = f"{cont_id} - {cliente} - {total:.2f}" if cliente else f"{cont_id} - {total:.2f}"
        nome = base
        k = 2
        while nome in usados:
            nome = f"{base} ({k})"
            k += 1
        usados.add(nome)
        blocos.append({
            "cont_id": str(cont_id),
            "cliente_nome": nome_cliente.get(cont_id) or "",
            "total": round(total, 2),
            "qtd_veiculos": len(linhas),
            "filename": f"{nome}.xlsx",
            "xlsx_bytes": xls,
        })
    return {"periodo": periodo, "blocos": blocos}


def build_lote_payload(blocos, periodo, email, empresa, opcao=""):
    """
    Monta o payload do LOTE para o webhook N8N.
    Cada .xlsx fica separado no array `lote`, em Base64.
    """
    import base64
    from datetime import datetime, date

    def iso(d):
        if isinstance(d, (datetime, date)):
            return d.strftime("%Y-%m-%d")
        return str(d) if d else None

    itens = []
    for b in blocos:
        itens.append({
            "cont_id": b["cont_id"],
            "cliente_nome": b["cliente_nome"],
            "filename": b["filename"],
            "mime_type": MIME_XLSX,
            "total": b["total"],
            "qtd_veiculos": b["qtd_veiculos"],
            "file_base64": base64.b64encode(b["xlsx_bytes"]).decode("ascii"),
        })

    return {
        "email": email.strip(),
        "empresa": empresa,
        "tipo": "detalhamento_medicao",
        "opcao": opcao,
        "periodo": {
            "data_inicial": iso(periodo[0]),
            "data_final": iso(periodo[1]),
            "qtd_dias": periodo[2],
        },
        "qtd_relatorios": len(itens),
        "lote": itens,
    }
